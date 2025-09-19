
from openpyxl import load_workbook, Workbook
EXCEL_PATH = "backend/storage.xlsx"

def get_latest_email():
    wb = load_workbook(EXCEL_PATH)
    sheet = wb.active
    for row in reversed(list(sheet.iter_rows(values_only=True))):
        email = row[2] if len(row) >= 3 else None
        name = row[1] if len(row) >= 2 else None
        if email and "@" in email and not email.lower().startswith("string"):
            print("✅ Found fallback email:", email)
            return email, name
    print("❌ No valid fallback email found")
    return None, None

print(get_latest_email())