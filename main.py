from fastapi import FastAPI
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from openpyxl import load_workbook, Workbook
from dotenv import load_dotenv
from datetime import datetime
from uuid import uuid4
import os
import smtplib
from email.message import EmailMessage
import requests
from chatbot.router import router as chatbot_router
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron import CronTrigger
from contextlib import asynccontextmanager


# Load environment variables
load_dotenv()
EMAIL_USER = os.getenv("EMAIL_USER")
EMAIL_PASS = os.getenv("EMAIL_PASS")
OPENROUTER_API_KEY = os.getenv("OPENROUTER_API_KEY")
AUTBOUND_API_KEY = os.getenv("AUTBOUND_API_KEY")
GITHUB_TOKEN=os.getenv("GITHUB_TOKEN")
DEBUG = True
# Initialize FastAPI
@asynccontextmanager
async def lifespan(app: FastAPI):
    # Startup logic (if needed)
    yield
    # Shutdown logic
    scheduler.shutdown()
app = FastAPI(lifespan=lifespan)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["https://axion-portfolio-ui-bqvq.vercel.app/","*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)
app.include_router(chatbot_router)
# Excel setup

EXCEL_PATH = "storage.xlsx"
if not os.path.exists(EXCEL_PATH):
    wb = Workbook()
    sheet = wb.active
    sheet.append([
        "id", "name", "email", "userType", "company", "answers", "status", "role",
        "timestamp", "subject", "body", "github", "linkedin", "source"
    ])
    wb.save(EXCEL_PATH)

# Models
class User(BaseModel):
    name: str
    email: str
    userType: str  # "hr" or "visitor"
    company: str = ""
    role: str = ""
    answers: str = ""  # Role description or "Not hiring" response

class ContactInfo(BaseModel):
    name: str
    github: str = ""
    linkedin: str = ""


# Static skills and projects
skills_str = (
    "FastAPI, Django, DeepFace, OpenCV, Redis, Celery, GitHub API, LinkedIn API, "
    "Emotion Recognition, WebSockets, React, TailwindCSS, Firebase, JWT, Docker"
)

top_projects = [
    {"name": "MindSync", "description": "Emotion-adaptive learning platform using DeepFace, OpenCV, and Django"},
    {"name": "Autism Support System", "description": "Real-time gesture recognition and emergency alerts for neurodiverse users"},
    {"name": "Claims API", "description": "Scalable backend for healthcare claims using FastAPI, Redis, and Celery"}
]

project_lines = "; ".join([f"{p['name']}: {p['description']}" for p in top_projects])

# Prompt builders
def build_role_aware_prompt(name, role, company, role_description):
    return f"""
You are a professional assistant writing on behalf of Mohammed Karab Ehtesham.

The recipient is {name}, a {role} at {company}.
They are hiring for: "{role_description}"

Mohammed’s skills: {skills_str}
Relevant projects: {project_lines}

Write a concise, emotionally intelligent email that compares Mohammed’s background to the role description. Highlight relevant skills and projects naturally. Mention the company name. Make the email feel human—like it was written by someone who genuinely admires the recipient’s work. Use natural phrasing, subtle warmth, and a conversational tone. Keep it concise—no more than 3 short paragraphs.

Return the subject and body separated by a newline.
"""

def build_future_opportunity_prompt(name, role, company):
    return f"""
You are a professional assistant writing on behalf of Mohammed Karab Ehtesham.

The recipient is {name}, a {role} at {company}.
They are not currently hiring.

Mohammed’s skills: {skills_str}
Relevant projects: {project_lines}

Write a warm, emotionally intelligent email that expresses admiration for the company’s work and invites future connection. Mention how Mohammed’s background aligns with their long-term vision. Make the email feel human—like it was written by someone who genuinely respects the recipient’s work. Use natural phrasing, subtle warmth, and a conversational tone. Keep it concise—no more than 3 short paragraphs.

Return the subject and body separated by a newline.
"""

# Email generation
def generate_email_from_prompt(prompt):
    headers = {
        "Authorization": f"Bearer {OPENROUTER_API_KEY}",
        "Content-Type": "application/json"
    }

    payload = {
        "model": "deepseek/deepseek-r1:free",
        "messages": [{"role": "user", "content": prompt}]
    }

    try:
        response = requests.post("https://openrouter.ai/api/v1/chat/completions", headers=headers, json=payload)
        result = response.json()
        content = result["choices"][0]["message"]["content"]

        lines = content.strip().split("\n")
        subject_line = next((line for line in lines if line.lower().startswith("subject:")), None)
        subject = subject_line.replace("Subject:", "").strip() if subject_line else "Let's stay connected"
        body_lines = [line for line in lines if not line.lower().startswith("subject:") and line.strip()]
        body = "\n".join(body_lines).strip()

        return subject, body

    except Exception as e:
        print("AI email generation failed:", e)
        return "Let's stay connected", "Hi, thank you for reaching out. I’ve attached my resume so we can continue the conversation."

# Email utility
def send_email(to_email, file_path_or_text, subject, body):
    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = EMAIL_USER
    msg["To"] = to_email
    msg.set_content(body)

    if os.path.exists(file_path_or_text):
        with open(file_path_or_text, "rb") as f:
            msg.add_attachment(
                f.read(),
                maintype="application",
                subtype="octet-stream",
                filename=os.path.basename(file_path_or_text)
            )

    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
        smtp.login(EMAIL_USER, EMAIL_PASS)
        smtp.send_message(msg)

# Log user and respond
@app.post("/log-visitor")
def log_user(data: User):
    wb = load_workbook(EXCEL_PATH)
    sheet = wb.active

    user_id = str(uuid4())
    hiring = not data.answers.lower().startswith("not hiring")

    if data.userType == "hr":
        if hiring:
            prompt = build_role_aware_prompt(data.name, data.role or "Hiring Manager", data.company, data.answers)
        else:
            prompt = build_future_opportunity_prompt(data.name, data.role or "Hiring Manager", data.company)

        subject, body = generate_email_from_prompt(prompt)
        send_email(data.email, "resume.pdf", subject, body)
    else:
        subject = ""
        body = ""

    sheet.append([
        user_id,
        data.name,
        data.email,
        data.userType,
        data.company,
        data.answers,
        f"{data.userType.capitalize()} Logged",
        data.role,
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        subject,
        body
    ])
    wb.save(EXCEL_PATH)

    return {"redirect": "https://axion-portfolio-ui-bqvq.vercel.app/"}

# Contact page outreach

# 🔍 Retrieve latest email from Excel# 🔍 Retrieve latest email from Excel
def get_latest_email():
    wb = load_workbook(EXCEL_PATH)
    sheet = wb.active
    for row in reversed(list(sheet.iter_rows(values_only=True))):
        email = row[2] if len(row) >= 3 else None
        name = row[1] if len(row) >= 2 else None
        if email and "@" in email and not email.lower().startswith("string"):
            print("✅ Found fallback email:", email)
            return email, name
    print("❌ No valid fallback email found")
    return None, None

# 📧 Send intro email
def send_email_contact(to_email, subject, body):
    try:
        msg = EmailMessage()
        msg["Subject"] = subject
        msg["From"] = EMAIL_USER
        msg["To"] = to_email
        msg.set_content(body)

        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
            smtp.login(EMAIL_USER, EMAIL_PASS)
            smtp.send_message(msg)

        if DEBUG:
            print(f"✅ Email sent to {to_email}")
    except Exception as e:
        print(f"❌ Email sending failed:", e)

# 📊 Log contact to Excel
def log_contact(name, email, github, linkedin, source="contact"):
    wb = load_workbook(EXCEL_PATH)
    sheet = wb.active
    updated = False
    EXPECTED_COLUMNS = 15

    for row in sheet.iter_rows(min_row=2):
        if len(row) >= 3 and row[2].value == email:
            row_index = row[0].row
            sheet.cell(row=row_index, column=13).value = f"https://github.com/{github}" if github else sheet.cell(row=row_index, column=13).value
            sheet.cell(row=row_index, column=14).value = f"https://linkedin.com/in/{linkedin}" if linkedin else sheet.cell(row=row_index, column=14).value
            sheet.cell(row=row_index, column=10).value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            sheet.cell(row=row_index, column=15).value = source
            updated = True
            break

    if not updated:
        new_row = [""] * EXPECTED_COLUMNS
        new_row[1] = name
        new_row[2] = email
        new_row[8] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        new_row[11] = f"https://github.com/{github}" if github else ""
        new_row[12] = f"https://linkedin.com/in/{linkedin}" if linkedin else ""
        new_row[14] = source
        sheet.append(new_row)

    wb.save(EXCEL_PATH)

def send_daily_excel():
    try:
        if os.path.exists(EXCEL_PATH):
            subject = f"Daily Visitor Log - {datetime.now().strftime('%d %b %Y')}"
            body = "Attached is the latest visitor log from Mohammed’s assistant system."
            send_email(EMAIL_USER, EXCEL_PATH, subject, body)
            if DEBUG:
                print("✅ Daily Excel sheet sent successfully.")
        else:
            print("❌ Excel file not found.")
    except Exception as e:
        print("❌ Failed to send daily Excel:", e)

# 🐙 GitHub follow
def follow_on_github(username):
    if not username:
        return
    url = f"https://api.github.com/user/following/{username}"
    headers = {
        "Authorization": f"token {GITHUB_TOKEN}",
        "Accept": "application/vnd.github+json"
    }
    try:
        response = requests.put(url, headers=headers, timeout=5)
        if response.status_code in [204, 304]:
            print(f"✅ Followed {username} on GitHub")
        else:
            print(f"❌ GitHub follow failed: {response.status_code} - {response.text}")
    except Exception as e:
        print(f"❌ GitHub follow error:", e)

# 🔗 LinkedIn outreach
def connect_on_linkedin(name, email):
    if not name:
        return
    headers = {
        "X-API-KEY": AUTBOUND_API_KEY,
        "Content-Type": "application/json"
    }
    contact_email = email if email else f"{name.lower().replace(' ', '')}@example.com"
    data = {
        "contactEmail": contact_email,
        "userEmail": EMAIL_USER,
        "contentType": "connectionRequest"
    }
    try:
        requests.post("https://api.autobound.ai/api/external/generate-content/v1", headers=headers, json=data, timeout=5)
        print(f"✅ LinkedIn outreach triggered for {name}")
    except Exception as e:
        print(f"❌ LinkedIn outreach failed:", e)

# 🚀 Main contact endpoint
@app.post("/contact-outreach")
def contact_outreach(info: ContactInfo):
    name = info.name.strip()
    github = info.github.strip()
    linkedin = info.linkedin.strip()

    # 📨 Always extract latest email from Excel
    email, name_from_excel = get_latest_email()
    if not email:
        print("❌ No valid email found in Excel")
        return {"status": "failed", "reason": "No email available"}
    if not name or name.lower() == "string":
        name = name_from_excel or "Visitor"

    # 📊 Log contact
    log_contact(name, email, github, linkedin)

    # 📨 Send intro email
    email_sent = False
    if not github and not linkedin:
        subject = "Excited to connect!"
        body = f"Hi {name}, thanks for reaching out! Mohammed’s portfolio is designed to engage, adapt, and respond with clarity and purpose. Whether you're exploring his work or looking to collaborate, you're always welcome here. Let’s stay connected."
        send_email_contact(email, subject, body)
        email_sent = True

    # 🐙 GitHub follow
    github_followed = False
    if github:
        follow_on_github(github)
        github_followed = True

    # 🔗 LinkedIn connect
    linkedin_connected = False
    if linkedin:
        connect_on_linkedin(name, email)
        linkedin_connected = True

    print("✅ Contact flow completed:", {
        "email_sent": email_sent,
        "github_followed": github_followed,
        "linkedin_connected": linkedin_connected
    })

    return {
        "status": "outreach triggered",
        "email_sent": email_sent,
        "github_followed": github_followed,
        "linkedin_connected": linkedin_connected
    }
scheduler = BackgroundScheduler()
scheduler.add_job(send_daily_excel, CronTrigger(hour=8, minute=0))  # Sends daily at 8:00 AM
scheduler.start()

@app.on_event("shutdown")
def shutdown_event():
    scheduler.shutdown()