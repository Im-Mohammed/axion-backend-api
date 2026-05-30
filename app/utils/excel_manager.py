"""
excel_manager.py
Primary visitor log store — reads/writes a local .xlsx file on the shared host.
Falls back gracefully if openpyxl is unavailable.
"""

import os
import logging
from datetime import datetime
from threading import Lock

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

logger = logging.getLogger("portfolio.excel")

HEADERS = [
    "ID", "Name", "Email", "UserType", "Company",
    "Answers", "Status", "Role", "Timestamp",
    "Subject", "Body", "GitHub", "LinkedIn",
    "ModelUsed", "Source", "IP",
]

# Header fill colour (dark navy)
HEADER_FILL  = "1F2D3D"
HEADER_FONT  = "FFFFFF"
ALT_ROW_FILL = "F2F5F9"


class ExcelManager:
    def __init__(self, filepath: str):
        self.filepath = filepath
        self._lock    = Lock()
        if OPENPYXL_OK:
            self._ensure_file()
        else:
            logger.warning("openpyxl not installed — Excel logging disabled.")

    # ── Internal helpers ────────────────────────────────────────────────
    def _ensure_file(self):
        if os.path.exists(self.filepath):
            return
        os.makedirs(os.path.dirname(self.filepath), exist_ok=True)
        wb = Workbook()
        ws = wb.active
        ws.title = "Visitors"
        self._write_headers(ws)
        wb.save(self.filepath)
        logger.info(f"Excel log created at {self.filepath}")

    def _write_headers(self, ws):
        fill   = PatternFill("solid", fgColor=HEADER_FILL)
        font   = Font(bold=True, color=HEADER_FONT, name="Calibri")
        border = Border(
            bottom=Side(style="thin", color="AAAAAA")
        )
        for col, header in enumerate(HEADERS, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill      = fill
            cell.font      = font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = border
        ws.row_dimensions[1].height = 22
        # Set column widths
        col_widths = [36, 22, 32, 12, 22, 40, 18, 22, 20, 40, 60, 30, 30, 30, 14, 18]
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    def _load_wb(self):
        return load_workbook(self.filepath)

    def _style_data_row(self, ws, row_idx: int):
        """Alternating row shading for readability."""
        if row_idx % 2 == 0:
            fill = PatternFill("solid", fgColor=ALT_ROW_FILL)
            for col in range(1, len(HEADERS) + 1):
                ws.cell(row=row_idx, column=col).fill = fill

    # ── Public API ──────────────────────────────────────────────────────
    def append_visitor(self, row: list):
        if not OPENPYXL_OK:
            logger.warning("Excel append skipped (openpyxl missing).")
            return
        with self._lock:
            try:
                wb = self._load_wb()
                ws = wb.active
                next_row = ws.max_row + 1
                for col, value in enumerate(row, start=1):
                    cell = ws.cell(row=next_row, column=col, value=value)
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                self._style_data_row(ws, next_row)
                wb.save(self.filepath)
                logger.info(f"Excel: row {next_row} written.")
            except Exception as e:
                logger.error(f"Excel append failed: {e}")

    def get_all_visitors(self) -> list[dict]:
        """Return all rows as list of dicts (for admin panel)."""
        if not OPENPYXL_OK:
            return []
        with self._lock:
            try:
                wb = self._load_wb()
                ws = wb.active
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    return []
                headers = [str(h) for h in rows[0]]
                return [
                    {headers[i]: row[i] for i in range(len(headers))}
                    for row in rows[1:]
                    if any(cell is not None for cell in row)
                ]
            except Exception as e:
                logger.error(f"Excel read failed: {e}")
                return []

    def get_latest_email(self) -> tuple[str | None, str | None]:
        """Return (email, name) of the most recent valid visitor."""
        records = self.get_all_visitors()
        for row in reversed(records):
            email = row.get("Email", "") or ""
            name  = row.get("Name",  "") or ""
            if email and "@" in email and not email.lower().startswith("string"):
                return email, name
        return None, None

    def update_contact(self, email: str, github: str, linkedin: str):
        """Update GitHub/LinkedIn for a visitor identified by email."""
        if not OPENPYXL_OK:
            return
        with self._lock:
            try:
                wb   = self._load_wb()
                ws   = wb.active
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    return
                header = [str(h) for h in rows[0]]
                email_col   = header.index("Email")   + 1
                github_col  = header.index("GitHub")  + 1
                linkedin_col = header.index("LinkedIn") + 1
                ts_col      = header.index("Timestamp") + 1

                for row_idx, row in enumerate(rows[1:], start=2):
                    if row[email_col - 1] == email:
                        if github:
                            ws.cell(row=row_idx, column=github_col).value  = f"https://github.com/{github}"
                        if linkedin:
                            ws.cell(row=row_idx, column=linkedin_col).value = f"https://linkedin.com/in/{linkedin}"
                        ws.cell(row=row_idx, column=ts_col).value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        break
                wb.save(self.filepath)
                logger.info(f"Excel: contact updated for {email}")
            except Exception as e:
                logger.error(f"Excel update_contact failed: {e}")

    def get_stats(self) -> dict:
        """Quick stats for the admin dashboard."""
        records = self.get_all_visitors()
        if not records:
            return {"total": 0, "hr": 0, "developer": 0, "visitor": 0, "today": 0}
        today = datetime.now().strftime("%Y-%m-%d")
        return {
            "total":     len(records),
            "hr":        sum(1 for r in records if r.get("UserType") == "hr"),
            "developer": sum(1 for r in records if r.get("UserType") == "developer"),
            "visitor":   sum(1 for r in records if r.get("UserType") == "visitor"),
            "today":     sum(1 for r in records if str(r.get("Timestamp", "")).startswith(today)),
        }