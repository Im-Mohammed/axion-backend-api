"""
admin_router.py
Secure admin API — JWT-protected JSON endpoints only.

Rate limiting:
  /login → strict (5 per 5 min) — brute force protection
  All other endpoints → protected by JWT, no extra rate limit needed

Data source:
  Primary  → Excel (/backend/logs/visitors.xlsx)
  Backup   → Google Sheets
  Strategy → merge both, deduplicate by ID, newest first
"""

import io
import logging
from datetime import datetime, timedelta, timezone

import bcrypt
import jwt
from fastapi import APIRouter, Depends, HTTPException, Request, Response
from fastapi.security import HTTPAuthorizationCredentials, HTTPBearer
from openpyxl import Workbook
from openpyxl.styles import Alignment, Color, Font, PatternFill
from pydantic import BaseModel

from app.services.sheets import get_all_from_sheets
from app.settings.config import get_settings
from app.utils.excel_manager import ExcelManager
from app.utils.rate_limiter import RateLimiter

logger   = logging.getLogger("portfolio.admin")
settings = get_settings()
router   = APIRouter(prefix="/admin", tags=["admin"])

_excel         = ExcelManager("/backend/logs/visitors.xlsx")
_login_limiter = RateLimiter.for_login()   # only login needs rate limiting

JWT_ALGORITHM    = "HS256"
JWT_EXPIRE_HOURS = 8
security         = HTTPBearer(auto_error=False)


# ── JWT ────────────────────────────────────────────────────────────────────
def create_token(username: str) -> str:
    payload = {
        "sub": username,
        "exp": datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRE_HOURS),
        "iat": datetime.now(timezone.utc),
    }
    return jwt.encode(payload, settings.jwt_secret, algorithm=JWT_ALGORITHM)


def verify_token(
    credentials: HTTPAuthorizationCredentials = Depends(security),
) -> str:
    if not credentials:
        raise HTTPException(status_code=401, detail="Not authenticated")
    try:
        payload = jwt.decode(
            credentials.credentials,
            settings.jwt_secret,
            algorithms=[JWT_ALGORITHM],
        )
        return payload["sub"]
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Session expired — please log in again")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")


# ── IP helper ──────────────────────────────────────────────────────────────
def _get_ip(request: Request) -> str:
    forwarded = request.headers.get("X-Forwarded-For", "")
    if forwarded:
        return forwarded.split(",")[0].strip()
    return request.client.host if request.client else "unknown"


# ── Merge helper ───────────────────────────────────────────────────────────
def _get_merged_records() -> list[dict]:
    """
    Merge Excel + Sheets, deduplicate by ID, Excel takes priority.
    Skipped visitors included — they have userType=skipped.
    """
    excel_records  = _excel.get_all_visitors()
    sheets_records = get_all_from_sheets()

    merged: dict[str, dict] = {}

    for record in excel_records:
        record_id = record.get("ID")
        if record_id:
            merged[record_id] = record

    for record in sheets_records:
        record_id = record.get("ID")
        if record_id and record_id not in merged:
            merged[record_id] = record

    records = list(merged.values())
    records.sort(key=lambda r: str(r.get("Timestamp", "")), reverse=True)
    return records


# ── Models ─────────────────────────────────────────────────────────────────
class LoginRequest(BaseModel):
    username: str
    password: str


# ── Login ──────────────────────────────────────────────────────────────────
@router.post("/login")
def admin_login(
    request: Request,
    data: LoginRequest,
    limiter: RateLimiter = Depends(lambda: _login_limiter),
):
    """
    Returns JWT on valid credentials.
    Rate limited — 5 attempts per 5 minutes per IP.
    """
    ip = _get_ip(request)
    if not limiter.is_allowed(ip):
        raise HTTPException(
            status_code=429,
            detail="Too many login attempts. Please wait 5 minutes."
        )

    if data.username != settings.admin_username:
        raise HTTPException(status_code=401, detail="Invalid credentials")

    if not settings.admin_password_hash:
        raise HTTPException(status_code=500, detail="Admin password not configured")

    try:
        valid = bcrypt.checkpw(
            data.password.encode("utf-8"),
            settings.admin_password_hash.encode("utf-8"),
        )
    except Exception as e:
        logger.error(f"bcrypt error: {e}")
        valid = False

    if not valid:
        logger.warning(f"Failed admin login for '{data.username}' from {ip}")
        raise HTTPException(status_code=401, detail="Invalid credentials")

    token = create_token(data.username)
    logger.info(f"Admin '{data.username}' logged in from {ip}.")
    return {"token": token, "expires_in_hours": JWT_EXPIRE_HOURS}


# ── Stats ──────────────────────────────────────────────────────────────────
@router.get("/stats")
def admin_stats(username: str = Depends(verify_token)):
    records = _get_merged_records()
    today   = datetime.now().strftime("%Y-%m-%d")
    return {
        "total":     len(records),
        "hr":        sum(1 for r in records if r.get("UserType") == "hr"),
        "developer": sum(1 for r in records if r.get("UserType") == "developer"),
        "visitor":   sum(1 for r in records if r.get("UserType") == "visitor"),
        "skipped":   sum(1 for r in records if r.get("UserType") == "skipped"),
        "today":     sum(1 for r in records if str(r.get("Timestamp", "")).startswith(today)),
    }


# ── Visitor list ───────────────────────────────────────────────────────────
@router.get("/visitors")
def admin_visitors(
    page: int = 1,
    per_page: int = 50,
    user_type: str = "",
    username: str = Depends(verify_token),
):
    records = _get_merged_records()

    if user_type:
        records = [r for r in records if r.get("UserType") == user_type]

    total     = len(records)
    start     = (page - 1) * per_page
    paginated = records[start : start + per_page]

    for r in paginated:
        if r.get("Body") and len(str(r["Body"])) > 120:
            r["Body"] = str(r["Body"])[:120] + "…"

    return {
        "total":    total,
        "page":     page,
        "per_page": per_page,
        "pages":    (total + per_page - 1) // per_page,
        "data":     paginated,
    }


# ── Single visitor ─────────────────────────────────────────────────────────
@router.get("/visitor/{visitor_id}")
def admin_visitor_detail(
    visitor_id: str,
    username: str = Depends(verify_token),
):
    records = _get_merged_records()
    record  = next((r for r in records if r.get("ID") == visitor_id), None)
    if not record:
        raise HTTPException(status_code=404, detail="Visitor not found")
    return record


# ── Export ─────────────────────────────────────────────────────────────────
@router.get("/export")
def admin_export(username: str = Depends(verify_token)):
    records = _get_merged_records()

    if not records:
        raise HTTPException(status_code=404, detail="No visitor data found")

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Visitors"

    headers = list(records[0].keys())
    fill    = PatternFill(fill_type="solid", fgColor=Color(rgb="FF1F2D3D"))
    font    = Font(bold=True, color="FFFFFF", name="Calibri")

    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
        cell           = ws.cell(row=1, column=col)
        cell.fill      = fill
        cell.font      = font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, record in enumerate(records, start=2):
        for col, key in enumerate(headers, start=1):
            ws.cell(row=row_idx, column=col, value=record.get(key, ""))

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"visitors_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return Response(
        content=buffer.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )